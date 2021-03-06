#! /usr/bin/python
'''
Replace the keywords in a word .docx template by the content
of the columns with the same names in a specific row(s) of an excel sheet.

syntax:  jreplace  xlsx  docx  rownumber1 rownumber2 ...

The keywords used in the word template must correspond exactly to columns
in the first row of the excel sheet. 

The first sheet in the excel document will be used.

The output file(s) will have the same name as the template, but with a
tag indicating the associated row number.

The docx template can be replaced by a directory tree with multiple docx
documents, in which case the search/replace will apply to all the documents, 
and a new directory tree will be created.

Carlos Allende Prieto, October 2016
Carlos, December 2016, fixed it to work with keywords within tables and 
                       handle page breaks
Carlos, December 2016, upgraded to work with dir trees containing templates
'''

import sys
import os
import shutil
import docx
import openpyxl
from docx.document import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import _Cell, Table
from docx.text.paragraph import Paragraph


#template='01.clausula.docx'
#rownum=2
#workbook='00. BBDD proyectos.xlsx'


#basic checks
def info():
    """
    inform users on syntax
    """
    print('syntax:  jreplace  xlsx  docx/dirtree  rownumber1 rownumber2 ...')
    return

def prechecks():
    """
    carry out basic checks on the command-line arguments
    """
    if len(sys.argv) < 4:
        print('ERROR: at least 3 arguments are required')
        info()
        exit()
    if int(sys.argv[3]) < 2:
        print('ERROR: the 3rd argument (rownum) must be >=2')
        info()
        exit()
    if sys.argv[1][-4:] != 'xlsx':
        print(sys.argv[1][-4:])
        print('ERROR: the 1st argument must be the name of the xlxs file')
        info()
        exit()

    if os.path.isfile(sys.argv[2]):
        if sys.argv[2][-4:] != 'docx':
            print('ERROR: the 2nd argument must be the name of the docx file')
            info()
            exit()
    elif os.path.isdir(sys.argv[2]):
        pass
    else: 
        print ("ERROR: template "+sys.argv[2]+" is neither a file nor a directory")
        info()
        exit()        

def iter_block_items(parent):
    """
    Yield each paragraph and table child within *parent*, in document order.
    Each returned value is an instance of either Table or Paragraph. *parent*
    would most commonly be a reference to a main Document object, but
    also works for a _Cell object, which itself can contain paragraphs and tables.
    """
    if isinstance(parent, Document):
        parent_elm = parent.element.body
    elif isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        raise ValueError("something's not right")

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)
            #pass

def jreplace_file(sheet,template,rownum,tag):
    """
    Does the replacement job on a single file (template) for a single row 
    in the (excel) sheet
    """
    #open word template
    doc = docx.Document(template)
    
    #find and replace
    #print("KEYWORDS:"),
    for colnum in range(1,sheet.max_column+1):
        i=0
        keyword=str(sheet.cell(row=1, column=colnum).value)
        target=str(sheet.cell(row=rownum, column=colnum).value)
        if ((keyword == None) | (target == None)): continue
        #print((keyword+'-').strip()),
        #sys.stdout.write((keyword+'-').strip())
        #sys.stdout.flush()
        #print ("%s-" % keyword.strip(),end='')
        for block in iter_block_items(doc): 
            if hasattr(block,'text'): #deals with page breaks
                if block.text == u'\n': continue
            if hasattr(block,'runs'): #if it has runs it is a paragraph
                inline = block.runs
                #print('we have a paragraph')
                if (len(inline) < 1): continue
                for j in range(len(inline)):
                    if inline[j].text == '': continue
                    text = inline[j].text.replace(keyword,target)
                    inline[j].text = text
            else:                    #if it doesn't, it is a table
                if not hasattr(block,'rows'): continue
                #print('we have a table')
                for row in block.rows: 
                    if not hasattr(row,'cells'): continue
                    for cell in row.cells:
                        if not hasattr(cell,'paragraphs'): continue
                        for paragraph in cell.paragraphs:
                            #print ("paragraph text=",paragraph.text)
                            inline=paragraph.runs
                            for j in range(len(inline)):
                                #print(i,j)
                                text = inline[j].text.replace(keyword,target)
                                inline[j].text = text

        i+=1

    #save output to a new doc
    outdoc=template[0:-5]+tag+'.docx'
    print('\n-- Writing output to '+outdoc)
    doc.save(outdoc)

def docker(arg,dirname,filenames):
    """
    Auxiliary function to getdocs, which picks docx files from dirs
    """
    #print dirname,filenames
    for file in filenames:
       fullname=dirname+'/'+file
       if (os.path.isfile(fullname) & (fullname[-4:]  == 'docx')): 
           arg.append(fullname)
           #print ('adding '+fullname)

def getdocxs(dir):
    """
    Gets all the docx files in a dir tree
    """
    files=[]
    #os.path.walk(dir,docker,files)
    for directory, dirnames, filenames in os.walk(dir):
      for file in filenames:
        fullname=dir+'/'+file
        if (os.path.isfile(fullname) & (fullname[-4:]  == 'docx')): 
          files.append(fullname)

    return files

def jreplace_dir(sheet,template,rownum,tag):
    """
    Does the replacement job on all docx documents in a tree dir (template) 
    for a single row in the (excel) sheet
    """
    print ('template is='+template)
    newdir=template+tag
    shutil.copytree(template,newdir)
    docfiles=getdocxs(newdir)
    print (docfiles)
    for file in docfiles:
        jreplace_file(sheet,file,rownum,'')
    

if __name__ == "__main__":

    #carry out basic checks
    prechecks()

    #assign parameters to variables
    workbook=sys.argv[1]
    template=sys.argv[2]
    rownumbers=sys.argv[3:]


    #open excel sheet
    wb = openpyxl.load_workbook(workbook)
    sheetnames=wb.get_sheet_names()
    sheet=wb.get_sheet_by_name(sheetnames[0])


    #loop over rownumbers
    for row in rownumbers:
        rownum=int(row)
  
        if rownum > sheet.max_row:
            print('ERROR: requested rownumber='+row+' but the excel sheet has only '+ \
                   str(sheet.max_row)+' rows') 
            continue

        tag=("_%04d" % rownum)

        if os.path.isfile(template):
            jreplace_file(sheet,template,rownum,tag)
        elif os.path.isdir(template):
            jreplace_dir(sheet,template,rownum,tag)
        else: 
            print ("ERROR: template "+template+" is neither a file nor a directory")
            break

